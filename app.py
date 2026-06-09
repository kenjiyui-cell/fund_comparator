import streamlit as st
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Fund Comparator", page_icon="📊", layout="wide",
                   initial_sidebar_state="expanded")

st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=IBM+Plex+Sans:wght@400;500;600&display=swap" rel="stylesheet">
<style>
:root{
  --sans:'IBM Plex Sans',system-ui,sans-serif;
  --mono:'IBM Plex Mono',ui-monospace,monospace;
  --bg:#f8f8f6;
  --surface:#ffffff;
  --surface-2:#f2f2ef;
  --line:#e5e5e0;
  --line-2:#d4d4ce;
  --ink-1:#1e2140;
  --ink-2:#4a4e6a;
  --ink-3:#8a8fa8;
  --accent:#4f6bed;
  --accent-bg:rgba(79,107,237,0.08);
  --ok:#1a9b6c;    --ok-bg:#eaf7f2;
  --bad:#d94f2a;   --bad-bg:#fdf0ec;
  --warn:#b07a10;  --warn-bg:#fdf5e0;
  --neutral-bg:#f0f0ec;
  --r:7px; --r-lg:11px;
  --shadow:0 1px 2px rgba(20,22,40,.04),0 1px 1px rgba(20,22,40,.03);
}
/* hide default streamlit chrome */
#MainMenu{visibility:hidden}
footer{visibility:hidden}
header{visibility:hidden}
.block-container{padding-top:0.5rem !important;padding-bottom:2rem !important;max-width:100% !important;}
[data-testid="stSidebar"]{background:var(--surface);border-right:1px solid var(--line);}
[data-testid="stSidebar"] .stMarkdown p{font-family:var(--sans);font-size:12px;color:var(--ink-2);}
.stButton>button{font-family:var(--sans) !important;border-radius:var(--r) !important;}
.stButton>button[kind="primary"]{background:var(--accent) !important;border:none !important;font-weight:600 !important;}
.stButton>button[kind="primary"]:hover{filter:brightness(1.07) !important;}
.stTextInput>div>div>input{font-family:var(--sans) !important;border-radius:var(--r) !important;font-size:12px !important;}
.stMultiSelect>div{border-radius:var(--r) !important;}
.stSelectbox>div>div{border-radius:var(--r) !important;}
.stCheckbox label{font-family:var(--sans) !important;font-size:13px !important;}
.stNumberInput>div>div>input{font-family:var(--mono) !important;font-size:12px !important;}
.stMetric{background:var(--surface);border:1px solid var(--line);border-radius:var(--r-lg);padding:13px 15px !important;box-shadow:var(--shadow);}
.stMetric label{font-size:11.5px !important;color:var(--ink-3) !important;font-family:var(--sans) !important;text-transform:uppercase;letter-spacing:.04em;}
.stMetric [data-testid="stMetricValue"]{font-size:27px !important;font-weight:600 !important;color:var(--ink-1) !important;letter-spacing:-.02em !important;font-family:var(--sans) !important;}
.stDataFrame{border:1px solid var(--line) !important;border-radius:var(--r-lg) !important;overflow:hidden !important;}
.stDataFrame thead th{background:var(--surface-2) !important;color:var(--ink-3) !important;font-size:10px !important;text-transform:uppercase !important;letter-spacing:.05em !important;font-family:var(--sans) !important;font-weight:600 !important;}
.stDataFrame tbody td{font-size:11.5px !important;font-family:var(--sans) !important;color:var(--ink-2) !important;}
.stExpander{border:1px solid var(--line) !important;border-radius:var(--r-lg) !important;}
.stExpander summary{font-family:var(--sans) !important;font-size:12.5px !important;font-weight:600 !important;}
div[data-testid="stVerticalBlock"]>div{gap:0.6rem;}
.stDivider{border-color:var(--line) !important;}
.stSubheader{font-family:var(--sans) !important;font-weight:600 !important;color:var(--ink-1) !important;letter-spacing:-.01em !important;}
.stWarning{border-radius:var(--r) !important;}
.stInfo{border-radius:var(--r) !important;}
.stSuccess{border-radius:var(--r) !important;}
/* app header */
.fc-topbar{display:flex;justify-content:space-between;align-items:center;font-size:11px;color:var(--ink-3);padding:6px 2px 10px;letter-spacing:.005em;font-family:var(--sans);}
.fc-topbar strong{color:var(--ink-1);font-weight:600;}
.fc-apphead{display:flex;align-items:center;gap:16px;background:var(--surface);border:1px solid var(--line);border-radius:var(--r-lg);padding:10px 16px;box-shadow:var(--shadow);margin-bottom:14px;}
.fc-brand{display:flex;align-items:center;gap:10px;}
.fc-mark{width:32px;height:32px;border-radius:8px;background:var(--ink-1);display:flex;align-items:center;justify-content:center;flex-shrink:0;}
.fc-mark svg{width:16px;height:16px;fill:var(--accent);}
.fc-name{font-size:15px;font-weight:600;letter-spacing:-.01em;color:var(--ink-1);line-height:1;font-family:var(--sans);}
.fc-name span{color:var(--accent);}
.fc-sub{font-size:10px;color:var(--ink-3);font-family:var(--mono);margin-top:2px;}
.fc-badge{background:var(--accent-bg);color:var(--accent);border-radius:5px;padding:3px 10px;font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.05em;font-family:var(--sans);margin-left:auto;}
/* sidebar cards */
.fc-card{background:var(--surface);border:1px solid var(--line);border-radius:var(--r-lg);box-shadow:var(--shadow);margin-bottom:10px;overflow:hidden;}
.fc-card-head{padding:8px 12px;border-bottom:1px solid var(--line);font-size:10px;font-weight:600;color:var(--ink-3);text-transform:uppercase;letter-spacing:.07em;font-family:var(--sans);}
.fc-card-body{padding:11px 12px;}
/* legend */
.fc-legend{display:flex;flex-direction:column;gap:4px;}
.fc-li{display:flex;align-items:center;gap:8px;font-size:11.5px;color:var(--ink-2);padding:4px 6px;border-radius:6px;font-family:var(--sans);}
.fc-dot{width:8px;height:8px;border-radius:50%;flex-shrink:0;}
/* section headers */
.fc-section{font-size:12px;font-weight:600;color:var(--ink-2);text-transform:uppercase;letter-spacing:.06em;font-family:var(--sans);margin:4px 0 8px;display:flex;align-items:center;gap:8px;}
.fc-section::after{content:"";flex:1;height:1px;background:var(--line);}
/* KPI stat cards (match the design mockup) */
.fc-stats{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:18px;}
.fc-stat{background:var(--surface);border:1px solid var(--line);border-radius:var(--r-lg);padding:16px 18px;box-shadow:var(--shadow);}
.fc-stat-val{font-size:30px;font-weight:600;line-height:1;letter-spacing:-.02em;font-family:var(--sans);}
.fc-stat-label{font-size:12px;color:var(--ink-3);margin-top:9px;font-family:var(--sans);}
@media(max-width:900px){.fc-stats{grid-template-columns:repeat(2,1fr);}}
/* results table polish */
.stDataFrame tbody tr:hover td{background:var(--surface-2) !important;}
.stCaption,.stCaption p{font-family:var(--mono) !important;font-size:11px !important;color:var(--ink-3) !important;}
</style>
""", unsafe_allow_html=True)

# ── App header ──────────────────────────────────────────────────────────────
st.markdown("""
<div class="fc-topbar">
  <span>SMDAM Fund Operations &nbsp;·&nbsp; Internal validation tool · Confidential</span>
  <strong>v3.1</strong>
</div>
<div class="fc-apphead">
  <div class="fc-brand">
    <div class="fc-mark">
      <svg viewBox="0 0 20 20" xmlns="http://www.w3.org/2000/svg">
        <rect x="1" y="11" width="4" height="8" rx="1"/>
        <rect x="8" y="6" width="4" height="13" rx="1"/>
        <rect x="15" y="1" width="4" height="18" rx="1"/>
      </svg>
    </div>
    <div>
      <div class="fc-name">Fund<span>Comparator</span></div>
      <div class="fc-sub">reconcile · validate · export</div>
    </div>
  </div>
  <span class="fc-badge">Compare</span>
</div>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════
ENTITY_REF = "entity reference"    # normalized form of the anchor column
ENTITY_NAME = "entity name"

# Columns that, by default, LINE UP rows (the matching key). Kept deliberately
# small so that everything else — including Currency*, Methodology*, Entity Type,
# Entity Name and all the value columns — gets COMPARED and can be flagged.
DEFAULT_KEY_COLS = {
    "entity reference", "value type", "as of date",
    "breakdown", "breakdown type",
}


def norm(s):
    """Normalize a header/sheet name: drop trailing *, collapse spaces, lowercase."""
    s = str(s).strip().rstrip("*").strip()
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def is_blank(v):
    return str(v).strip().lower() in ("", "nan", "none", "nat")


def key_norm(v):
    """Normalize a cell VALUE for use inside a matching key (handles dates)."""
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            pass
    return str(v).strip().lower()


def find_col(df, normalized_target):
    for c in df.columns:
        if norm(c) == normalized_target:
            return c
    return None


def compare_value(big_v, small_v, tol):
    try:
        return "match" if abs(float(big_v) - float(small_v)) <= tol else "mismatch"
    except (ValueError, TypeError):
        return "match" if str(big_v).strip() == str(small_v).strip() else "mismatch"


def detect_header_row(raw_df, max_scan=10):
    """Find the row (0-based) that contains 'Entity Reference' by scanning the top rows."""
    for i in range(min(max_scan, len(raw_df))):
        for cell in raw_df.iloc[i].tolist():
            if norm(cell) == ENTITY_REF:
                return i
    return None


def _resolve_header_idx(data, sheet, header_row):
    """Return the 0-based header index for one sheet. header_row=None => auto-detect."""
    if header_row is not None:
        return header_row - 1
    try:
        raw = pd.read_excel(io.BytesIO(data), sheet_name=sheet, header=None, nrows=12)
    except Exception:
        return None
    return detect_header_row(raw)


@st.cache_data(show_spinner=False)
def get_big_columns(data: bytes, header_row):
    """Read just the headers of every sheet to populate the key-column picker.
    header_row=None means auto-detect the Entity Reference* row per sheet."""
    xls = pd.ExcelFile(io.BytesIO(data))
    seen, cols = set(), []
    for sn in xls.sheet_names:
        hidx = _resolve_header_idx(data, sn, header_row)
        if hidx is None:
            continue
        try:
            hdr = pd.read_excel(io.BytesIO(data), sheet_name=sn, header=hidx, nrows=0)
        except Exception:
            continue
        for c in hdr.columns:
            c = str(c).strip()
            if c and norm(c) not in seen:
                seen.add(norm(c))
                cols.append(c)
    return cols


@st.cache_data(show_spinner=False)
def get_big_sheet_names(data: bytes):
    """List every sheet name in the workbook (for the sheet picker)."""
    return pd.ExcelFile(io.BytesIO(data)).sheet_names


def read_workbook(data: bytes, header_row, only_sheets=None):
    """Read sheets into DataFrames. header_row=None => auto-detect per sheet.
    Returns (sheets_dict, info_dict) where info explains what happened per sheet."""
    xls = pd.ExcelFile(io.BytesIO(data))
    sheets, info = {}, {}
    for name in xls.sheet_names:
        if only_sheets is not None and name not in only_sheets:
            continue
        hidx = _resolve_header_idx(data, name, header_row)
        if hidx is None:
            info[name] = "no Entity Reference* header found — skipped"
            continue
        try:
            df = pd.read_excel(io.BytesIO(data), sheet_name=name, header=hidx)
        except Exception as e:
            info[name] = f"could not read ({e}) — skipped"
            continue
        df.columns = [str(c).strip() for c in df.columns]
        sheets[name] = df
        info[name] = f"header on row {hidx + 1}, {len(df)} rows"
    return sheets, info


# ════════════════════════════════════════════════════════════════
#  THE COMPARISON ENGINE
# ════════════════════════════════════════════════════════════════
def run_comparison(big_sheets, small_files, tol, key_cols_norm):
    # Entity Reference* is always part of the key.
    key_cols_norm = list(dict.fromkeys([ENTITY_REF] + [k for k in key_cols_norm if k != ENTITY_REF]))

    norm_big_sheet = {norm(name): name for name in big_sheets}
    results, empty_headers, notes, diag = [], [], [], []

    # 1) Empty headers in the big file (per sheet)
    for sname, bdf in big_sheets.items():
        for col in bdf.columns:
            if len(bdf) == 0 or bdf[col].apply(is_blank).all():
                empty_headers.append({"Sheet": sname, "Header": str(col).rstrip("*").strip()})

    # 2) Compare each fund file against the matching big sheet
    for fname, ssheets in small_files:
        for s_sheet, sdf in ssheets.items():
            big_name = norm_big_sheet.get(norm(s_sheet))
            if big_name is None:
                notes.append(f"⚠️ {fname}: sheet “{s_sheet}” has no matching sheet in the big file — skipped.")
                diag.append({"Fund file": fname, "Sheet": s_sheet, "Big rows": "—",
                             "Small rows": len(sdf), "Matched rows": 0,
                             "Outcome": "no matching sheet name in big file"})
                continue

            bdf = big_sheets[big_name]
            big_n2c = {norm(c): c for c in bdf.columns}
            small_n2c = {norm(c): c for c in sdf.columns}

            # Key columns that actually exist in BOTH files for this sheet
            active_key = [k for k in key_cols_norm if k in big_n2c and k in small_n2c]
            if ENTITY_REF not in active_key:
                where = "big" if ENTITY_REF not in big_n2c else "small"
                notes.append(f"⚠️ {fname}/{s_sheet}: no Entity Reference* column in the {where} file — skipped.")
                diag.append({"Fund file": fname, "Sheet": s_sheet, "Big rows": len(bdf),
                             "Small rows": len(sdf), "Matched rows": 0,
                             "Outcome": f"no Entity Reference* column in {where} file"})
                continue

            # Build big lookup keyed on the composite key
            big_lookup, dups = {}, 0
            for _, brow in bdf.iterrows():
                k = tuple(key_norm(brow[big_n2c[c]]) for c in active_key)
                if k in big_lookup:
                    dups += 1
                    continue
                big_lookup[k] = brow
            if dups:
                notes.append(f"ℹ️ Big file/{big_name}: {dups} row(s) share an identical key "
                             f"({', '.join(active_key)}); only the first of each was kept.")

            nonempty_big = {c for c in bdf.columns if len(bdf) > 0 and not bdf[c].apply(is_blank).all()}
            matched_rows = 0

            for _, srow in sdf.iterrows():
                ent_ref = str(srow[small_n2c[ENTITY_REF]]).strip()
                if is_blank(ent_ref):
                    continue

                k = tuple(key_norm(srow[small_n2c[c]]) for c in active_key)
                brow = big_lookup.get(k)
                in_big = brow is not None
                if in_big:
                    matched_rows += 1

                # context fields
                if ENTITY_NAME in big_n2c and in_big:
                    ent_name = str(brow[big_n2c[ENTITY_NAME]])
                elif ENTITY_NAME in small_n2c:
                    ent_name = str(srow[small_n2c[ENTITY_NAME]])
                else:
                    ent_name = ""
                id_parts = [str(srow[small_n2c[c]]) for c in active_key if c not in (ENTITY_REF, ENTITY_NAME)]
                row_identity = " | ".join(p for p in id_parts if not is_blank(p))

                if not in_big:
                    # report the whole row once, not per column
                    results.append({
                        "Status": "row_not_found", "Sheet": big_name,
                        "Entity Reference*": ent_ref, "Entity Name": ent_name,
                        "Row identity": row_identity, "Header": "(entire row)",
                        "Big (correct)": "", "Small (export)": "", "Fund file": fname,
                    })
                    continue

                for s_col in sdf.columns:
                    ncol = norm(s_col)
                    if ncol in active_key:
                        continue                       # key columns are used to match, not compared
                    big_col = big_n2c.get(ncol)
                    if big_col is None or big_col not in nonempty_big:
                        continue                       # not in big, or empty header
                    sv = srow[s_col]
                    bv = brow[big_col]
                    status = "missing" if is_blank(sv) else compare_value(bv, sv, tol)
                    results.append({
                        "Status": status, "Sheet": big_name,
                        "Entity Reference*": ent_ref, "Entity Name": ent_name,
                        "Row identity": row_identity, "Header": str(big_col).rstrip("*").strip(),
                        "Big (correct)": bv, "Small (export)": sv, "Fund file": fname,
                    })

            outcome = "compared OK" if matched_rows else "0 rows lined up — check key columns / values"
            diag.append({"Fund file": fname, "Sheet": big_name, "Big rows": len(bdf),
                         "Small rows": len(sdf), "Matched rows": matched_rows, "Outcome": outcome})

    results_df = pd.DataFrame(results)
    empty_df = (pd.DataFrame(empty_headers).drop_duplicates()
                if empty_headers else pd.DataFrame(columns=["Sheet", "Header"]))
    diag_df = pd.DataFrame(diag) if diag else pd.DataFrame(
        columns=["Fund file", "Sheet", "Big rows", "Small rows", "Matched rows", "Outcome"])
    return results_df, empty_df, notes, diag_df


# ════════════════════════════════════════════════════════════════
#  STEP 1 — UPLOAD
# ════════════════════════════════════════════════════════════════
st.markdown('<div class="fc-section">Step 1 — Upload files</div>', unsafe_allow_html=True)
col1, col2 = st.columns(2)
with col1:
    st.markdown("**Big file** — the correct reference (multiple funds, multiple sheets)")
    big_file = st.file_uploader("Big file", type=["xlsx"], key="big", label_visibility="collapsed")
with col2:
    st.markdown("**Small files** — one fund each (website exports). Drop in as many as you like.")
    small_uploads = st.file_uploader("Small files", type=["xlsx"], key="small",
                                     accept_multiple_files=True, label_visibility="collapsed")

# ── Settings ──
st.markdown('<div class="fc-section">Step 2 — Settings</div>', unsafe_allow_html=True)
auto_header = st.checkbox("Auto-detect the header row on each sheet (recommended)", value=True,
                          help="Finds the row containing Entity Reference* on every sheet automatically, "
                               "so sheets with headers on different rows all work. Untick to set rows manually.")
s1, s2, s3 = st.columns(3)
with s1:
    big_header_row = st.number_input("Big file header row", 1, 10, 2, disabled=auto_header,
                                     help="Used only when auto-detect is off.")
with s2:
    small_header_row = st.number_input("Small file header row", 1, 10, 1, disabled=auto_header,
                                       help="Used only when auto-detect is off.")
with s3:
    tolerance = st.selectbox("Numeric tolerance", [0, 0.01, 0.1, 1.0], index=1,
                             format_func=lambda x: "Exact match only" if x == 0 else f"±{x} (ignore rounding)")

big_hdr = None if auto_header else big_header_row
small_hdr = None if auto_header else small_header_row

key_cols_display = []
sheets_to_use = None
if big_file:
    big_bytes = big_file.getvalue()

    # Let the user pick which sheets to compare (defaults to the known data sheets).
    KNOWN_DATA_SHEETS = {"valuation", "valuation period", "return", "gain loss",
                         "payment", "pa contribution"}
    all_sheet_names = get_big_sheet_names(big_bytes)
    default_sheets = [s for s in all_sheet_names if norm(s) in KNOWN_DATA_SHEETS] or all_sheet_names
    sheets_to_use = st.multiselect(
        "Worksheets to compare",
        options=all_sheet_names, default=default_sheets,
        help="Pick the data sheets. Leave out cover pages or instruction tabs "
             "(e.g. 提出方法), which don't follow the header layout.",
    )

    all_cols = get_big_columns(big_bytes, big_hdr)
    default_key = [c for c in all_cols if norm(c) in DEFAULT_KEY_COLS]
    key_cols_display = st.multiselect(
        "Columns used to LINE UP rows (not compared)",
        options=all_cols, default=default_key,
        help="The tool pairs a big row with a small row when ALL of these match — keep this set small "
             "(just the columns that tell one of a fund's rows from another). EVERY other column is "
             "compared and flagged. Entity Reference* is always included.",
    )

# ════════════════════════════════════════════════════════════════
#  STEP 3 — RUN
# ════════════════════════════════════════════════════════════════
if big_file and small_uploads:
    st.divider()
    if st.button("▶  Compare files", type="primary", use_container_width=True):
        try:
            big_sheets, big_info = read_workbook(big_bytes, big_hdr, only_sheets=sheets_to_use)
            small_files, small_info = [], {}
            for f in small_uploads:
                sheets, info = read_workbook(f.getvalue(), small_hdr)
                small_files.append((f.name, sheets))
                small_info[f.name] = info
            key_cols_norm = [norm(c) for c in key_cols_display]
            results_df, empty_df, notes, diag_df = run_comparison(
                big_sheets, small_files, tolerance, key_cols_norm)
            st.session_state["results_df"] = results_df
            st.session_state["empty_df"] = empty_df
            st.session_state["notes"] = notes
            st.session_state["diag_df"] = diag_df
            st.session_state["big_info"] = big_info
            st.session_state["small_info"] = small_info
        except Exception as e:
            st.error(f"Something went wrong while reading the files: {e}")


# ════════════════════════════════════════════════════════════════
#  RESULTS
# ════════════════════════════════════════════════════════════════
if "results_df" in st.session_state:
    results_df = st.session_state["results_df"]
    empty_df = st.session_state["empty_df"]
    notes = st.session_state["notes"]
    diag_df = st.session_state.get("diag_df")
    big_info = st.session_state.get("big_info", {})
    small_info = st.session_state.get("small_info", {})

    for n in notes:
        st.warning(n)

    # Per-sheet diagnostics — shows exactly what happened on each sheet
    with st.expander("🔧 Per-sheet diagnostics (open this if a sheet isn't comparing)", expanded=results_df.empty):
        st.markdown("**How each sheet was read**")
        big_rows = [{"Sheet": k, "Result": v} for k, v in big_info.items()]
        st.caption("Big file:")
        st.dataframe(pd.DataFrame(big_rows) if big_rows else pd.DataFrame(columns=["Sheet", "Result"]),
                     use_container_width=True, hide_index=True)
        for fname, info in small_info.items():
            st.caption(f"Small file — {fname}:")
            rows = [{"Sheet": k, "Result": v} for k, v in info.items()]
            st.dataframe(pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Sheet", "Result"]),
                         use_container_width=True, hide_index=True)
        if diag_df is not None and not diag_df.empty:
            st.markdown("**Matching outcome per sheet**")
            st.dataframe(diag_df, use_container_width=True, hide_index=True)

    st.divider()
    st.markdown('<div class="fc-section">Results</div>', unsafe_allow_html=True)

    if results_df.empty:
        st.info("No comparable rows were found. Open the diagnostics above — it shows, per sheet, whether the "
                "header was found, how many rows each side had, and how many lined up.")
    else:
        total = len(results_df)
        matched = (results_df["Status"] == "match").sum()
        mismatch = (results_df["Status"] == "mismatch").sum()
        missing = (results_df["Status"] == "missing").sum()
        not_found = (results_df["Status"] == "row_not_found").sum()
        match_rate = round(100 * matched / total) if total else 0

        # Polished KPI cards (matches the Claude design mockup)
        st.markdown(f"""
<div class="fc-stats">
  <div class="fc-stat">
    <div class="fc-stat-val" style="color:var(--accent)">{total:,}</div>
    <div class="fc-stat-label">Cells compared</div>
  </div>
  <div class="fc-stat">
    <div class="fc-stat-val" style="color:var(--ok)">{match_rate}%</div>
    <div class="fc-stat-label">Match rate</div>
  </div>
  <div class="fc-stat">
    <div class="fc-stat-val" style="color:var(--bad)">{mismatch:,}</div>
    <div class="fc-stat-label">Mismatches</div>
  </div>
  <div class="fc-stat">
    <div class="fc-stat-val" style="color:var(--warn)">{missing + not_found:,}</div>
    <div class="fc-stat-label">Missing / not found</div>
  </div>
</div>
""", unsafe_allow_html=True)

        status_labels = {"match": "✅ Match", "mismatch": "🔴 Mismatch",
                         "missing": "🟡 No value (small)", "row_not_found": "⚪ Row not in big"}

        st.sidebar.header("🔎 Filters (Excel-style)")
        st.sidebar.markdown("""
<div class="fc-card">
  <div class="fc-card-head">Legend</div>
  <div class="fc-card-body">
    <div class="fc-legend">
      <div class="fc-li"><div class="fc-dot" style="background:var(--ok)"></div>Match — values are identical</div>
      <div class="fc-li"><div class="fc-dot" style="background:var(--bad)"></div>Mismatch — values differ</div>
      <div class="fc-li"><div class="fc-dot" style="background:var(--warn)"></div>No value in small file</div>
      <div class="fc-li"><div class="fc-dot" style="background:var(--ink-3)"></div>Row not found in big file</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)
        def msel(label, col):
            opts = sorted(results_df[col].astype(str).unique())
            return st.sidebar.multiselect(label, opts, default=opts)
        sheets_sel = msel("Sheet", "Sheet")
        funds_sel = msel("Entity Reference* (fund)", "Entity Reference*")
        headers_sel = msel("Header", "Header")
        status_sel = st.sidebar.multiselect("Status", list(status_labels.keys()),
                                            default=list(status_labels.keys()),
                                            format_func=lambda s: status_labels.get(s, s))
        search = st.sidebar.text_input("Search (any text)")

        view = results_df[
            results_df["Sheet"].astype(str).isin(sheets_sel)
            & results_df["Entity Reference*"].astype(str).isin(funds_sel)
            & results_df["Header"].astype(str).isin(headers_sel)
            & results_df["Status"].isin(status_sel)
        ]
        if search:
            mask = view.apply(lambda r: search.lower() in " ".join(map(str, r.values)).lower(), axis=1)
            view = view[mask]

        def style_table(df):
            styles = pd.DataFrame("", index=df.index, columns=df.columns)
            colors = {
                "match": "background-color: #c8e6c9; color: #1b5e20",
                "mismatch": "background-color: #ffcdd2; color: #b71c1c",
                "missing": "background-color: #fff9c4; color: #f57f17",
                "row_not_found": "background-color: #eeeeee; color: #555555",
            }
            for i, row in df.iterrows():
                c = colors.get(row["_raw"], "")
                for col in ["Big (correct)", "Small (export)"]:
                    if col in df.columns:
                        styles.at[i, col] = c
            return styles

        disp = view.copy()
        disp["_raw"] = disp["Status"]
        disp["Status"] = disp["Status"].map(status_labels).fillna(disp["Status"])
        ordered = ["Status", "Sheet", "Entity Reference*", "Entity Name", "Row identity",
                   "Header", "Big (correct)", "Small (export)", "Fund file", "_raw"]
        disp = disp[[c for c in ordered if c in disp.columns]]

        styled = disp.style.apply(style_table, axis=None).hide(axis="columns", subset=["_raw"])
        st.dataframe(styled, use_container_width=True, height=520)
        st.caption(f"Showing {len(view):,} of {total:,} checks · "
                   "🟢 match · 🔴 mismatch · 🟡 no value in small · ⚪ row not found in big")

    # ── Empty headers ──
    st.divider()
    st.markdown('<div class="fc-section">Headers with no value in the big file</div>', unsafe_allow_html=True)
    if empty_df.empty:
        st.success("Every header in the big file has at least one value.")
    else:
        st.caption("These columns are empty in the big file, so there is nothing to compare against.")
        st.dataframe(empty_df, use_container_width=True, height=240)

    # ── Export ──
    st.divider()
    st.markdown('<div class="fc-section">Export results</div>', unsafe_allow_html=True)

    def build_excel(comp_df, empty_df):
        wb = Workbook(); ws = wb.active; ws.title = "Comparison"
        fills = {"match": PatternFill("solid", fgColor="C8E6C9"),
                 "mismatch": PatternFill("solid", fgColor="FFCDD2"),
                 "missing": PatternFill("solid", fgColor="FFF9C4"),
                 "row_not_found": PatternFill("solid", fgColor="EEEEEE")}
        fonts = {"match": Font(bold=True, color="1B5E20"),
                 "mismatch": Font(bold=True, color="B71C1C"),
                 "missing": Font(bold=True, color="F57F17"),
                 "row_not_found": Font(bold=True, color="555555")}
        hdr_fill, hdr_font = PatternFill("solid", fgColor="212121"), Font(bold=True, color="FFFFFF")
        thin = Border(*(Side(style="thin", color="DDDDDD"),) * 4)

        cols = ["Status", "Sheet", "Entity Reference*", "Entity Name", "Row identity",
                "Header", "Big (correct)", "Small (export)", "Fund file"]
        cols = [c for c in cols if c in comp_df.columns]
        ws.append(cols)
        for cell in ws[1]:
            cell.fill, cell.font, cell.border = hdr_fill, hdr_font, thin
            cell.alignment = Alignment(horizontal="center")

        bi = cols.index("Big (correct)") if "Big (correct)" in cols else -1
        si = cols.index("Small (export)") if "Small (export)" in cols else -1
        for _, row in comp_df.iterrows():
            ws.append([row[c] for c in cols])
            r = ws.max_row
            key = str(row.get("Status", ""))
            f, ft = fills.get(key), fonts.get(key)
            for ci, cell in enumerate(ws[r], start=1):
                cell.border, cell.alignment = thin, Alignment(horizontal="left")
                if ci - 1 in (bi, si) and f:
                    cell.fill, cell.font = f, ft
        for ci, cells in enumerate(ws.columns, start=1):
            w = max((len(str(c.value or "")) for c in cells), default=10)
            ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 60)

        ws2 = wb.create_sheet("Empty headers")
        ws2.append(["Sheet", "Header"])
        for cell in ws2[1]:
            cell.fill, cell.font = hdr_fill, hdr_font
        for _, row in empty_df.iterrows():
            ws2.append([row["Sheet"], row["Header"]])
        for ci, cells in enumerate(ws2.columns, start=1):
            w = max((len(str(c.value or "")) for c in cells), default=10)
            ws2.column_dimensions[get_column_letter(ci)].width = min(w + 4, 50)

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf

    src = results_df if not results_df.empty else pd.DataFrame(
        columns=["Status", "Sheet", "Entity Reference*", "Entity Name", "Row identity",
                 "Header", "Big (correct)", "Small (export)", "Fund file"])
    st.download_button("⬇  Download Excel report (.xlsx)",
                       data=build_excel(src, empty_df),
                       file_name="fund_comparison_report.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True, type="primary")
